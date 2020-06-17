#!/usr/bin/env python
# coding: utf-8


                                    ## -----Importing required modules----- ##

import pandas as pd
import numpy as np
from pathlib import Path
import os
from pprint import pprint
import sys
from openpyxl import load_workbook
import logging


'''
Set the path for the log file here, this file will contain all error logged during runtime
'''
logging.basicConfig(filename='/Users/amankumaryadav/Data Science/Back up/cleaner.log',
                    filemode='w', level=logging.INFO,
                    format='%(asctime)s %(levelname)s: %(message)s',
                    datefmt='%m/%d/%Y %I:%M:%S %p')


##--------------------------------------------------------------------------------------------------------


                                         ## -----Data Cleaning----- ##


# Class containing functions to import template file and data files and
# write dataframes to excel files
class handle_data_from_files():


    # Initialise file address with default directory to be used
    def __init__(self, address = "/Users/amankumaryadav/Data Science/Back up/"):

        self.directory = Path(address)


    # Function to import template data
    '''
    Input :  Takes filename as input
    Output : Returns dataframe containing data from the template file
    '''
    def importTemplate(self, templateFilename = "Template.xlsx"):

        try:

            # Input path for file
            inputFilePath = self.directory

            templateFile = inputFilePath / templateFilename
            templateData = pd.read_excel(templateFile, sheet_name= 0)
            templateData.fillna(0, inplace = True)

            return templateData

        except FileNotFoundError:

            print(sys.exc_info()[:2])



    # Function to import file
    '''
    Input :  Takes filename as input
    Output : Returns dictionary where key is the sheet name and value is the dataframe
             containing sheet data
    '''
    def importFile(self, filename, header = 0):

        try:

            # Input path for file
            inputFilePath = self.directory

            inputFile = inputFilePath / filename
            inputData = pd.read_excel(inputFile, sheet_name= None, header = header)

            return inputData

        except FileNotFoundError:

            print(sys.exc_info()[:2])





    # Function to write dictionary of dataframes to an new excel file
    '''
    Input: Dictionary to be written to new excel file, where key is the sheet
           name and value is the corresponding table
    Output: None
    '''

    def writeDictToExcel(self, dictionary, filename):

        if filename in os.listdir(self.directory):

            outputFile = self.directory / filename

            with pd.ExcelWriter(outputFile, engine="openpyxl", mode = "a") as writer:

                for key in dictionary.keys():

                    df = dictionary[key]

                    if not df.empty:

                        df.to_excel(writer, sheet_name=key, index = False)

                    else:

                        pass

        else:

            outputFile = self.directory / filename

            with pd.ExcelWriter(outputFile) as writer:

                for key in dictionary.keys():

                    df = dictionary[key]

                    if not df.empty:

                        df.to_excel(writer, sheet_name=key, index = False)

                    else:

                        pass


##--------------------------------------------------------------------------------------------------------


# Class containing functions to clean dataframes
class clean_data():

    '''
    Contains the headers for the output file, can be changed as per user
    requirement. Just pass the new list below
    '''
    def __init__(self):

        self.names = ["Purchase Order", "Invoice", "Invoice Date",
                      "Invoice Value", "Payee Number"]



    # Function to rename column names of a dataframe
    '''
    Input: Dataframe and list of new names
    Output: Dataframe with new column names
    '''
    def renameDataframe(self, df, names = None):

        if names == None:
            new_names = dict(zip(df.columns, self.names))
            df.rename(columns = new_names, inplace = True)
        else:
            new_names = dict(zip(df.columns, names))
            df.rename(columns = new_names, inplace = True)

        return df



    # Function to convert template data to dictionary
    '''
    Input: Dataframe, name of keys to be used in dict
    Output: Dictionary with column name as key and corresponding values as list
    '''
    def cleanDict(self, df_dict):

        for key in df_dict.keys():

            df_dict[key] = [val for val in df_dict[key] if val != 0]

        return df_dict


##--------------------------------------------------------------------------------------------------------


# Class containing function to get clean dataframes from the imported workbook
class clean_workbook_data():


    # Function to process sheet data
    '''
    Input: Sheet data as a dataframe from a file
    Output: Dictionary where key is the sheet name and value is the dataframe
    '''

    def processSheet(self, sheetData, headers, sheetName):

        '''
        Sets dataframe to appropriate value in case of improper
        dataframe header.
        '''
        if "Unnamed: 0" in sheetData.columns and (not sheetData.empty):

            for i in range(10):

                if all(isinstance(x,str) for x in sheetData.iloc[i]):
                    sheetData.columns = sheetData.iloc[i]
                    sheetData = sheetData.iloc[i+1:]
                    break
                else:
                    continue



        '''
        Checks if the dataframe has a proper header name or
        not after processing by above if statement, else skips
        the dataframe processing.
        '''
        if "Unnamed: 0" not in sheetData.columns and (not sheetData.empty):


            df = pd.DataFrame()
            columns = list(sheetData.columns)
            dueDateFlag = False
            poFlag = False
            data_dict = dict.fromkeys(headers,0)
            finalData = {}


            # Sort values by due date
            for value in headers['Invoice Date']:

                if value in columns:

                    df = sheetData.sort_values(by = value, ascending = True)
                    dueDateFlag = True

                    break


            # Looks for purchase order column
            for value in headers['Purchase Order']:

                if value in columns:

                    poFlag = True
                    poName = value
                    df[poName] = df[poName].astype(str)

                    break


            # Check if purchase order column exists in the input data
            if poFlag:


                df[poName] = df[poName].apply(lambda a: a.split('.',1)[0])

                checkCondition = (df[poName].str.startswith('2') & df[poName].str.isdigit())

                dfFiltered = df[checkCondition]
                dfException = df[~checkCondition]

                finalData[f"{sheetName[:15]}_filtered"] = dfFiltered
                finalData[f"{sheetName[:15]}_exception"] = dfException

                for data, name in [(dfFiltered, "Filtered")]:

                    newData = pd.DataFrame()
                    data_dict = dict.fromkeys(headerMapping,0)
                    filteredColumns = data.columns

                    for key in headers.keys():

                        values = headers[key]

                        for value in values:

                            if ((value in filteredColumns) and (data_dict[key] < 1)):


                                if (sheetData[value].empty):

                                    newData[key] = np.nan

                                else:

                                    newData[key] = data[value]
                                    data_dict[key] += 1

                            elif (value not in filteredColumns) and data_dict[key] < 1:

                                newData[key] = np.nan

                    finalData[f"{sheetName[:10]}_{name}_full"] = newData

                return finalData


            else:

                newData = pd.DataFrame()
                data_dict = dict.fromkeys(headerMapping,0)
                filteredColumns = sheetData.columns

                for key in headers.keys():

                    values = headers[key]

                    for value in values:

                        if ((value in filteredColumns) and (data_dict[key] < 1)):


                            if (sheetData[value].empty):

                                newData[key] = np.nan

                            else:

                                newData[key] = sheetData[value]
                                data_dict[key] += 1

                        elif (value not in filteredColumns) and data_dict[key] < 1:

                            newData[key] = np.nan

                finalData[f"{sheetName[:10]}_SortedByDate"] = df
                finalData[f"{sheetName[:10]}_SortedByDatefull"] = newData

                return finalData




    # Function to clean all the sheet in a workbook
    '''
    Input: Excel workbook
    Output: Dictionary containing cleaned data for all the sheets in the workbook
    '''

    def cleanWorkbook(self, workbookData, headers):

        workbookDict = {}

        for key in list(workbookData.keys()):

            df = workbookData[key]
            cleanedData = self.processSheet(df, headers, key)
            workbookDict[key] = cleanedData

        return workbookDict


##--------------------------------------------------------------------------------------------------------

# Specify the template file below. It will be used to create header mapping where keys are headers
# to be put in the final sheet and values are a list of all possible synonyms of the header.
# Import the template data file

try:

    file = handle_data_from_files()
    templateData = file.importTemplate()

except:

    logging.info(sys.exc_info()[:2])
    print(sys.exc_info()[:2])

##--------------------------------------------------------------------------------------------------------

# headerMapping is a dictionary that contains the mapping of columns names to the header required
# in the final output. Here key is the column name in the output and values is the list of all
# possible columns names in the input data.
# Getting the header mappings
'''
Function cleanDict from class clean_data takes a template dictionary and
cleans is before transfroming it into header mapping.
'''

try:

    dataCleaner = clean_data()
    templateData = dataCleaner.renameDataframe(templateData)
    templateDict = templateData.to_dict(orient = "list")
    headerMapping = dataCleaner.cleanDict(templateDict)

except:

    logging.info(sys.exc_info()[:2])
    print(sys.exc_info()[:2])


##--------------------------------------------------------------------------------------------------------


# Specify the excel file below. In case of error the error is logged in an error log file which
# shares the same directory as the excel files.
'''
Importing data file by specifying file name, the same file
will be used to write the cleaned data.
'''
# Specify filename to be imported

# Uncomment this to input filename on terminal
# dataFileName = str(input("Enter the input file name:\n"))

dataFileName = str(input("Enter the file name: "))

try:

    sampleData = file.importFile(dataFileName + ".xlsx")

except:

    # File not found error
    logging.info(sys.exc_info()[:2])
    print(sys.exc_info()[:2])

##--------------------------------------------------------------------------------------------------------


# Below code writes the data from the input file into a new file.
# Write input data in a new file
# Uncomment this to input filename on terminal
# outputFile = str(input("Enter the output file name:\n"))

outputFile = f"{dataFileName} output.xlsx"
file.writeDictToExcel(sampleData, outputFile)


##--------------------------------------------------------------------------------------------------------


# Below code cleans the workbook and returns a dictionary where key is the sheet name and value is
# a dictionary of new sheets as key which is to be written in the file, with dataframes as their values.
# Cleaning the workbook
try:

    book = clean_workbook_data()
    workBook = book.cleanWorkbook(sampleData, headerMapping)

except:

    logging.info(sys.exc_info()[:2])
    print(sys.exc_info()[:2])


##--------------------------------------------------------------------------------------------------------



# The final dataframe is written in the same file from which the data was imported, though the user
# can specify the file of one's choice.
try:

    for sheet in workBook.keys():

        if workBook[sheet] != None:

            file.writeDictToExcel(workBook[sheet], outputFile)

        else:

            pass

except:

    logging.error(sys.exc_info()[:2])
    print(sys.exc_info()[:2])